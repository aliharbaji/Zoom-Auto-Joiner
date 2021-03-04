import webbrowser
from datetime import datetime
from time import sleep
import pyautogui
import openpyxl

try:
    import pyautogui
    import webbrowser
    import openpyxl
    from datetime import datetime
    import time

except ModuleNotFoundError as err:
    print("libraries not installed, please go through the README file, press enter to exit")
    input()
    exit()

t = 5  # sleep time (minutes) this helps the program rest for a couple minutes.
t = t * 60

#   *** gathering data from data sheet *** checks if the information written in the excel files are legit or not

Monday = openpyxl.load_workbook('Daily Program/0-Monday.xlsx')
mondaySheet = Monday.active

Tuesday = openpyxl.load_workbook('Daily Program/1-Tuesday.xlsx')
tuesdaySheet = Tuesday.active

Wednesday = openpyxl.load_workbook('Daily Program/2-Wednesday.xlsx')
wednesdaySheet = Wednesday.active

Thursday = openpyxl.load_workbook('Daily Program/3-Thursday.xlsx')
thursdaySheet = Thursday.active

Friday = openpyxl.load_workbook('Daily Program/4-Friday.xlsx')
fridaySheet = Friday.active

mondayArray = []
tuesdayArray = []
wednesdayArray = []
thursdayArray = []
fridayArray = []

for i in range(2, mondaySheet.max_row + 1):
    isActualClass = True
    for j in range(1, 3):
        if str(mondaySheet.cell(row=i, column=j).value).lower() == 'none':
            isActualClass = False
    if not isActualClass:
        continue

    arr = [str(mondaySheet.cell(row=i, column=1).value).strip()[:-3],
           str(mondaySheet.cell(row=i, column=2).value).strip()[:-3],
           str(mondaySheet.cell(row=i, column=3).value).strip(), str(mondaySheet.cell(row=i, column=4).value).strip(),
           "0"]
    mondayArray.append(arr)

for i in range(2, tuesdaySheet.max_row + 1):
    isActualClass = True
    for j in range(1, 3):
        if str(tuesdaySheet.cell(row=i, column=j).value).lower() == 'none':
            isActualClass = False
    if not isActualClass:
        continue

    arr = [str(tuesdaySheet.cell(row=i, column=1).value).strip()[:-3],
           str(tuesdaySheet.cell(row=i, column=2).value).strip()[:-3],
           str(tuesdaySheet.cell(row=i, column=3).value).strip(), str(tuesdaySheet.cell(row=i, column=4).value).strip(),
           "0"]
    tuesdayArray.append(arr)

for i in range(2, wednesdaySheet.max_row + 1):
    isActualClass = True
    for j in range(1, 3):
        if str(wednesdaySheet.cell(row=i, column=j).value).lower() == 'none':
            isActualClass = False
    if not isActualClass:
        continue

    arr = [str(wednesdaySheet.cell(row=i, column=1).value).strip()[:-3],
           str(wednesdaySheet.cell(row=i, column=2).value).strip()[:-3],
           str(wednesdaySheet.cell(row=i, column=3).value).strip(),
           str(wednesdaySheet.cell(row=i, column=4).value).strip(), "0"]
    wednesdayArray.append(arr)

for i in range(2, thursdaySheet.max_row + 1):
    isActualClass = True
    for j in range(1, 3):
        if str(thursdaySheet.cell(row=i, column=j).value).lower() == 'none':
            isActualClass = False
    if not isActualClass:
        continue

    arr = [str(thursdaySheet.cell(row=i, column=1).value).strip()[:-3],
           str(thursdaySheet.cell(row=i, column=2).value).strip()[:-3],
           str(thursdaySheet.cell(row=i, column=3).value).strip(),
           str(thursdaySheet.cell(row=i, column=4).value).strip(), "0"]
    thursdayArray.append(arr)

for i in range(2, fridaySheet.max_row + 1):
    isActualClass = True
    for j in range(1, 3):
        if str(fridaySheet.cell(row=i, column=j).value).lower() == 'none':
            isActualClass = False
    if not isActualClass:
        continue

    arr = [str(fridaySheet.cell(row=i, column=1).value).strip()[:-3],
           str(fridaySheet.cell(row=i, column=2).value).strip()[:-3],
           str(fridaySheet.cell(row=i, column=3).value).strip(), str(fridaySheet.cell(row=i, column=4).value).strip(),
           "0"]
    fridayArray.append(arr)

week = [mondayArray, tuesdayArray, wednesdayArray, thursdayArray, fridayArray]

#   *** finished gathering data from data sheet ***


on = True
isStarted = False

while on:

    datetime = datetime.now()
    time = datetime.strftime("%H:%M:%S")
    today = datetime.now()
    weekday = today.weekday()

    hour = datetime.now().hour
    minute = datetime.now().minute

    print(time, weekday)  # keeps showing you the time

    # 0 Monday, 1 Tuesday, 2 Wednesday, 3 Thursday, 4 Friday, 5 Saturday, 6 Sunday

    # this if statement checks if you have school today, if not the program stops
    if weekday > 4:
        print("You don't have school today, smartypants!")
        sleep(5)
        exit()

    for i in range(len(week[weekday])):  # foreach class of your classes

        if not isStarted:

            if hour == int(week[weekday][i][0].split(":")[0]) and minute + t / 60 >= int(week[weekday][i][0].split(":")[1]) and week[weekday][i][4] == "0":
                webbrowser.open("https://edu-il.zoom.us/j/" + week[weekday][i][2])  # opens the Zoom link
                sleep(10)  # waits 10 seconds and let's the Zoom window load
                pyautogui.typewrite(week[weekday][i][3])  # types meeting's password
                sleep(0.2)
                pyautogui.press("enter")  # clicks enter
                print("You joined the class.")
                week[weekday][i][4] = "1"

                isStarted = True  # tells the program that you joined the class so it begins to check whether it is
                # time to leave it or not

        if isStarted:
            if hour == int(week[weekday][i][1].split(":")[0]) and minute >= int(week[weekday][i][1].split(":")[1]) and \
                    week[weekday][i][4] == "1":
                pyautogui.hotkey("alt", 'x')
                print("You left the class.")
                isStarted = False  # tells the program that you left the class so it begins to check whether it is
                # time to join another one or not

        if hour > int(week[weekday][len(week[weekday]) - 1][1].split(":")[0]) or (hour == int(week[weekday][len(week[weekday]) - 1][1].split(":")[0]) and minute >= int(week[weekday][len(week[weekday]) - 1][1].split(":")[1])):  # checks if the program already joined the last class
            # or the class is over

            print("The program finished joining all the classes, Goodbye!")
            sleep(5)  # waits 5 seconds before eliminating
            on = False
            exit()

    if not on:  # check if the program is still working, if not: exit
        exit()

    sleep(t)  # makes the program stop running for a few minutes
